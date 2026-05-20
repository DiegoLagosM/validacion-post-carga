import os
import extract_msg
from openpyxl import load_workbook
from datetime import datetime

# -------------------------------------------------
# CONFIGURACIÓN
# -------------------------------------------------
RUTA_CORREOS = r"C:\Users\c31352e\OneDrive - EXPERIAN SERVICES CORP\Archivos de Cardenas, Oscar - Calidad de Datos\Diego\Validacion post carga\Correos"

RUTA_EXCEL = r"C:\Users\c31352e\OneDrive - EXPERIAN SERVICES CORP\Archivos de Cardenas, Oscar - Calidad de Datos\Diego\Validacion post carga\Consolidado validaciones de carga 2026.xlsx"

HOJA_EXCEL = "R04"

ASUNTO_CORREO = "Estadistica Proceso DSF - w_TransformadorDSF_R04"

# Campos que vienen en el correo
CAMPOS = {
    "Total de Registros": "total",
    "Total de Registros Actualizados": "actualizados",
    "Total de Registros Erroneos": "erroneos",
    "Porcentaje de Erroneos": "porcentaje"
}

# Filas FIJAS por campo (no cambian nunca)
FILAS_CAMPO = {
    "Total de Registros": 2,
    "Total de Registros Actualizados": 3,
    "Total de Registros Erroneos": 4,
    "Porcentaje de Erroneos": 5
}

# -------------------------------------------------
# FUNCIONES
# -------------------------------------------------
def obtener_o_crear_columna_fecha(ws, fecha):
    for col in range(2, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == fecha:
            return col

    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = fecha
    return col


def obtener_fecha_msg(msg):
    if isinstance(msg.date, datetime):
        return msg.date.strftime("%Y-%m-%d")

    fecha_raw = msg.date.split("+")[0].strip()
    return datetime.strptime(
        fecha_raw, "%a, %d %b %Y %H:%M:%S"
    ).strftime("%Y-%m-%d")


def extraer_valor(linea, concepto):
    if concepto.lower() not in linea.lower():
        return None

    valor = linea[len(concepto):].strip()
    valor = valor.lstrip(" :-=.•\t")
    return valor if valor else None


def leer_correo(ruta):
    for archivo in os.listdir(ruta):
        if not archivo.lower().endswith(".msg"):
            continue

        msg = extract_msg.Message(os.path.join(ruta, archivo))
        subject = (msg.subject or "").lower()

        if ASUNTO_CORREO.lower() in subject:
            data = {"fecha": obtener_fecha_msg(msg)}

            for linea in msg.body.splitlines():
                for texto, clave in CAMPOS.items():
                    valor = extraer_valor(linea, texto)
                    if valor:
                        data[clave] = valor

            msg.close()
            return data

        msg.close()

    return None


# -------------------------------------------------
# PROCESO PRINCIPAL
# -------------------------------------------------
wb = load_workbook(RUTA_EXCEL)

# Crear hoja si no existe
if HOJA_EXCEL in wb.sheetnames:
    ws = wb[HOJA_EXCEL]
else:
    ws = wb.create_sheet(HOJA_EXCEL)
    ws.cell(row=1, column=1).value = "Concepto"
    for campo, fila in FILAS_CAMPO.items():
        ws.cell(row=fila, column=1).value = campo
    print(f"⚠️ Hoja '{HOJA_EXCEL}' creada automáticamente.")

correo = leer_correo(RUTA_CORREOS)

if not correo:
    raise ValueError("❌ No se encontró el correo R04")

col_fecha = obtener_o_crear_columna_fecha(ws, correo["fecha"])

# Escritura de valores (UPSERT real)
for campo, clave in CAMPOS.items():
    fila = FILAS_CAMPO[campo]
    ws.cell(row=fila, column=col_fecha).value = correo.get(clave, "NO INFORMADO")

# -------------------------------------------------
# GUARDAR
# -------------------------------------------------
wb.save(RUTA_EXCEL)
print("✅ Proceso R04 ejecutado correctamente")