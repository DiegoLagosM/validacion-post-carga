from pathlib import Path
import re
import sys
import pandas as pd
from openpyxl import load_workbook

# =========================================================
# CONFIGURACIÓN
# =========================================================
BASE_PATH = Path(
    r"C:/Users/c31352e/OneDrive - EXPERIAN SERVICES CORP/Archivos de Cardenas, Oscar - Calidad de Datos/Diego/Validacion post carga"
)

RUTA_LOGS = BASE_PATH / "LOGS"
RUTA_EXCEL = BASE_PATH / "Consolidado validaciones de carga 2026.xlsx"

HOJA_CONSOLIDADO = "CCP-ONP"

# Regex precompilados
NAME_RE = re.compile(
    r'^AA([A-Z]+)(\d+)(20\d{6})1\.log$',
    re.IGNORECASE
)

CONTENT_RE = re.compile(
    r'Numero\s+de\s+Registros\s+Correctos\s+Agregados\s*:\s*(\d+)',
    re.IGNORECASE
)

# =========================================================
# FUNCIONES
# =========================================================
def obtener_logs_validos(path: Path) -> pd.DataFrame:
    registros = []

    if not path.exists():
        print(f"No existe la ruta de logs: {path}")
        return pd.DataFrame()

    for archivo in path.glob("*.log"):
        match = NAME_RE.match(archivo.name)
        if not match:
            continue

        tipo, codigo, fecha_raw = match.groups()

        try:
            fecha = (
                pd.to_datetime(fecha_raw, format="%Y%m%d")
                .strftime("%Y-%m-%d")
            )
        except ValueError:
            continue

        valor = None
        try:
            with archivo.open(encoding="utf-8", errors="ignore") as f:
                for linea in f:
                    m = CONTENT_RE.search(linea)
                    if m:
                        valor = int(m.group(1))
                        break
        except OSError:
            continue

        registros.append((codigo, tipo, fecha, valor))

    return pd.DataFrame(
        registros,
        columns=["CodigoBanco", "Tipo", "Fecha", "Valor"]
    )


def cargar_excel(path: Path, hoja: str) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=["CodigoBanco", "Tipo"])

    try:
        wb = load_workbook(path, read_only=True)
        if hoja not in wb.sheetnames:
            return pd.DataFrame(columns=["CodigoBanco", "Tipo"])
    except Exception:
        return pd.DataFrame(columns=["CodigoBanco", "Tipo"])

    return pd.read_excel(path, sheet_name=hoja, dtype=str)


def actualizar_datos(existente: pd.DataFrame, nuevo: pd.DataFrame) -> pd.DataFrame:
    if existente.empty:
        return nuevo.copy()

    existente = existente.set_index(["CodigoBanco", "Tipo"])
    nuevo = nuevo.set_index(["CodigoBanco", "Tipo"])

    existente = existente.reindex(existente.index.union(nuevo.index))

    # Alinear tipos
    for col in existente.columns.intersection(nuevo.columns):
        try:
            nuevo[col] = nuevo[col].astype(existente[col].dtype)
        except Exception:
            pass

    existente.update(nuevo)

    return existente.reset_index()


def ordenar_columnas(df: pd.DataFrame) -> pd.DataFrame:
    columnas_fijas = ["CodigoBanco", "Tipo"]
    columnas_fecha = sorted(
        c for c in df.columns
        if isinstance(c, str) and re.fullmatch(r"\d{4}-\d{2}-\d{2}", c)
    )
    return df[columnas_fijas + columnas_fecha]


# =========================================================
# EJECUCIÓN PRINCIPAL
# =========================================================
df_logs = obtener_logs_validos(RUTA_LOGS)

if df_logs.empty:
    print("No se encontraron logs válidos.")
    sys.exit(0)

nuevo = (
    df_logs
    .pivot(index=["CodigoBanco", "Tipo"], columns="Fecha", values="Valor")
    .reset_index()
)

existente = cargar_excel(RUTA_EXCEL, HOJA_CONSOLIDADO)

combinado = actualizar_datos(existente, nuevo)

combinado.sort_values(["CodigoBanco", "Tipo"], inplace=True)
combinado = ordenar_columnas(combinado)

with pd.ExcelWriter(
    RUTA_EXCEL,
    engine="openpyxl",
    mode="a" if RUTA_EXCEL.exists() else "w",
    if_sheet_exists="replace"
) as writer:
    combinado.to_excel(writer, sheet_name=HOJA_CONSOLIDADO, index=False)

print("✅ Proceso completado correctamente")
