import os
import extract_msg
from openpyxl import load_workbook
from datetime import datetime
import unicodedata
import re

# -------------------------------------------------
# CONFIGURACIÓN
# -------------------------------------------------
RUTA_CORREOS = r"C:\Users\c31352e\OneDrive - EXPERIAN SERVICES CORP\Archivos de Cardenas, Oscar - Calidad de Datos\Diego\Validacion post carga\Correos"

RUTA_EXCEL = r"C:\Users\c31352e\OneDrive - EXPERIAN SERVICES CORP\Archivos de Cardenas, Oscar - Calidad de Datos\Diego\Validacion post carga\Consolidado validaciones de carga 2026.xlsx"

HOJA_EXCEL = "IG - CA"

ASUNTOS = [
    "Estadisticas Proceso Cargador Cambio de Genero",
    "Estadisticas Proceso Cargador Cambio Orden Apellidos"
]

BLOQUES = {
    "Estadisticas Proceso Cargador Cambio de Genero": {
        "titulo": "Cambio de Género",
        "inicio": 2
    },
    "Estadisticas Proceso Cargador Cambio Orden Apellidos": {
        "titulo": "Cambio Orden Apellidos",
        "inicio": 6
    }
}

# -------------------------------------------------
# UTILIDADES
# -------------------------------------------------
def normalizar_texto(texto):
    if not texto:
        return ""
    texto = texto.lower()
    texto = unicodedata.normalize('NFD', texto)
    texto = texto.encode('ascii', 'ignore').decode("utf-8")
    return texto


def limpiar_numero(valor):
    if not valor:
        return ""
    return re.sub(r"[^\d]", "", valor)


# -------------------------------------------------
# EXCEL
# -------------------------------------------------
def obtener_o_crear_columna_fecha(ws, fecha):
    for col in range(2, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == fecha:
            return col

    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = fecha
    return col


# -------------------------------------------------
# EXTRACCIÓN
# -------------------------------------------------
def obtener_fecha_msg(msg):
    try:
        if isinstance(msg.date, datetime):
            return msg.date.strftime("%Y-%m-%d")

        fecha_raw = msg.date.split("+")[0].strip()
        return datetime.strptime(
            fecha_raw, "%a, %d %b %Y %H:%M:%S"
        ).strftime("%Y-%m-%d")
    except:
        return "SIN_FECHA"


# ✅ FUNCIÓN CLAVE CORREGIDA
def extraer_valores(linea):
    resultado = {}
    linea_norm = normalizar_texto(linea)

    # ✅ Archivo (ARSBXXXX)
    match_archivo = re.search(r"\b[A-Z]{4}\d{4}\b", linea)
    if match_archivo:
        resultado["archivo_procesado"] = match_archivo.group(0)

    # ✅ SOLO línea correcta de cantidad total origen
    if ("cantidad" in linea_norm and 
        "total" in linea_norm and 
        "origen" in linea_norm):

        numero = re.search(r"\b\d{2,}\b", linea)
        if numero:
            resultado["cantidad_total"] = numero.group(0)

    return resultado


def leer_correos(ruta):
    if not os.path.exists(ruta):
        print(f"❌ ERROR: Ruta no encontrada -> {ruta}")
        return {}

    datos = {}

    for archivo in os.listdir(ruta):
        if not archivo.lower().endswith(".msg"):
            continue

        path_msg = os.path.join(ruta, archivo)

        try:
            msg = extract_msg.Message(path_msg)
            subject = msg.subject or ""
            body = msg.body or ""

            for asunto in ASUNTOS:
                if normalizar_texto(asunto) in normalizar_texto(subject):

                    resultado = {
                        "fecha": obtener_fecha_msg(msg)
                    }

                    for linea in body.splitlines():
                        valores = extraer_valores(linea)

                        # ✅ IMPORTANTE: no sobrescribir valores correctos
                        for k, v in valores.items():
                            if k not in resultado:
                                resultado[k] = v

                    datos.setdefault(asunto, []).append(resultado)

            msg.close()

        except Exception as e:
            print(f"Error leyendo {archivo}: {e}")
            continue

    return datos


# -------------------------------------------------
# ESCRITURA EXCEL
# -------------------------------------------------
def escribir_excel(datos):
    if not os.path.exists(RUTA_EXCEL):
        print(f"❌ ERROR: No existe el Excel -> {RUTA_EXCEL}")
        return

    wb = load_workbook(RUTA_EXCEL)

    if HOJA_EXCEL in wb.sheetnames:
        ws = wb[HOJA_EXCEL]
    else:
        ws = wb.create_sheet(HOJA_EXCEL)
        ws.cell(row=1, column=1).value = "Concepto"

    for asunto, lista in datos.items():

        bloque = BLOQUES.get(asunto)
        if not bloque:
            continue

        for data in lista:

            fila = bloque["inicio"]
            col_fecha = obtener_o_crear_columna_fecha(ws, data["fecha"])

            ws.cell(row=fila, column=1).value = bloque["titulo"]

            # Archivo
            ws.cell(row=fila + 1, column=1).value = "Archivo Procesado"
            ws.cell(row=fila + 1, column=col_fecha).value = data.get(
                "archivo_procesado", "NO INFORMADO"
            )

            # Cantidad correcta
            total = limpiar_numero(data.get("cantidad_total", ""))

            ws.cell(row=fila + 2, column=1).value = "Cantidad Total en Archivo Origen"
            ws.cell(row=fila + 2, column=col_fecha).value = (
                int(total) if total.isdigit() else "NO INFORMADO"
            )

    wb.save(RUTA_EXCEL)
    print("✅ Excel actualizado correctamente")


# -------------------------------------------------
# MAIN
# -------------------------------------------------
if __name__ == "__main__":
    print("🚀 Iniciando proceso...")

    correos = leer_correos(RUTA_CORREOS)

    if correos:
        escribir_excel(correos)
    else:
        print("⚠️ No se encontraron datos")

    print("🏁 Proceso finalizado")