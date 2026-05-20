import os
import extract_msg
from openpyxl import load_workbook
from datetime import datetime
import re

# -------------------------------------------------
# CONFIGURACIÓN
# -------------------------------------------------
RUTA_CORREOS = r"C:\Users\c31352e\OneDrive - EXPERIAN SERVICES CORP\Archivos de Cardenas, Oscar - Calidad de Datos\Diego\Validacion post carga\Correos"

RUTA_EXCEL = r"C:\Users\c31352e\OneDrive - EXPERIAN SERVICES CORP\Archivos de Cardenas, Oscar - Calidad de Datos\Diego\Validacion post carga\Consolidado validaciones de carga 2026.xlsx"

HOJA_EXCEL = "SIR"

FILTRO_ASUNTO = "boletin concursal"

CAMPOS = {
    "0999BCJE": "bcje",
    "0999BCJP": "bcjp",
    "0999BCAP": "bcap"
}

FILAS_CAMPO = {
    "0999BCJE": 2,
    "0999BCJP": 3,
    "0999BCAP": 4
}

# -------------------------------------------------
# FUNCIONES
# -------------------------------------------------

def obtener_o_crear_columna_fecha(ws, fecha):
    for col in range(2, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == fecha:
            return col

    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = fecha
    return col


def obtener_fecha_msg(msg):
    if isinstance(msg.date, datetime):
        return msg.date.strftime("%Y-%m-%d")

    try:
        fecha_raw = msg.date.split("+")[0].strip()
        return datetime.strptime(
            fecha_raw, "%a, %d %b %Y %H:%M:%S"
        ).strftime("%Y-%m-%d")
    except:
        return "SIN_FECHA"


def extraer_valores_linea(linea):
    """
    Detecta diferentes formatos como:
    0999BCJE202604231 = 1500
    0999BCJE202604231 : 1500
    0999BCJE202604231 - 1500
    """

    patron = r"(0999[A-Z]{4})\d+\D+(\d+)"
    matches = re.findall(patron, linea, re.IGNORECASE)

    resultados = {}
    for codigo, valor in matches:
        resultados[codigo.upper()] = valor

    return resultados


def leer_correos(ruta):
    resultados = []

    for archivo in os.listdir(ruta):
        if not archivo.lower().endswith(".msg"):
            continue

        path = os.path.join(ruta, archivo)
        msg = extract_msg.Message(path)

        subject = (msg.subject or "").lower()

        print("-----")
        print(f"📧 Archivo: {archivo}")
        print(f"Asunto: {subject}")

        if FILTRO_ASUNTO in subject:

            data = {"fecha": obtener_fecha_msg(msg)}
            print("✅ CORREO VÁLIDO")

            for linea in msg.body.splitlines():
                print("LINEA:", linea)

                valores = extraer_valores_linea(linea)

                for codigo, valor in valores.items():
                    if codigo in CAMPOS:
                        print(f"✔ Detectado {codigo} -> {valor}")
                        data[CAMPOS[codigo]] = valor

            resultados.append(data)

        else:
            print("⛔ Ignorado por asunto")

        msg.close()

    return resultados


# -------------------------------------------------
# PROCESO PRINCIPAL
# -------------------------------------------------

wb = load_workbook(RUTA_EXCEL)

if HOJA_EXCEL in wb.sheetnames:
    ws = wb[HOJA_EXCEL]
else:
    ws = wb.create_sheet(HOJA_EXCEL)
    ws.cell(row=1, column=1).value = "Concepto"

    for campo, fila in FILAS_CAMPO.items():
        ws.cell(row=fila, column=1).value = campo

    print(f"⚠️ Hoja '{HOJA_EXCEL}' creada automáticamente.")

correos = leer_correos(RUTA_CORREOS)

if not correos:
    raise ValueError("❌ No se encontraron correos válidos")

for correo in correos:
    col_fecha = obtener_o_crear_columna_fecha(ws, correo["fecha"])

    for campo, clave in CAMPOS.items():
        fila = FILAS_CAMPO[campo]
        ws.cell(row=fila, column=col_fecha).value = correo.get(clave, "NO INFORMADO")

# -------------------------------------------------
# GUARDAR
# -------------------------------------------------

wb.save(RUTA_EXCEL)

print("✅ Proceso ejecutado correctamente")